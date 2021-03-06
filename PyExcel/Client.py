from  openpyxl import Workbook
import openpyxl
import os

class Client(object):
    def __init__(self, wb, sheet=None):
        if not wb.endswith(".xlsx"):
            wb = "%s.xlsx" %wb
        if os.path.exists(wb):
            self.wb = openpyxl.load_workbook(wb)
        else:
            self.wb = Workbook()
        if sheet in self.wb.sheetnames:
            self.sheet = self.wb[sheet]
        else:
            self.sheet = self.wb.create_sheet(sheet, 0)
        self.rows = list(self.sheet.iter_rows())
        self.head_row = list(self.rows)[0] if self.rows else []

    def save(self, file):
        if not file.endswith(".xlsx"):
            file = "%s.xlsx" % file
        self.wb.save(file)

    def insert_row(self, stringrow):
        row=stringrow.split(',')
        self.sheet.append(row)

    def get_head(self):
        return self.head_row

    def get_all(self):
        return list(self.rows)

    def get_col(self, num):
        datas = list(self.rows)
        col = []
        for data in datas:
            col.append(data[num])
        return col

    def get_sheet_names(self):
        return [ sheetname for sheetname in self.wb.sheetnames if not
        sheetname.startswith("Sheet")]

    def get_first_column_names(self):
        row_names = []
        max_row = self.sheet.max_row
        for row_num in range(max_row):
            row_names.append(self.sheet.cell(row_num+1, 1).value)
        return row_names

    def update_row_by_rownum(self, row_num, stringrow):
        row = stringrow.split(",")
        max_col = self.sheet.max_column
        if len(row) == max_col:
            for col_id in range(max_col):
                self.sheet.cell(int(row_num), col_id+1, row[col_id])
        else:
            print("row length is %s and max column is %s" %(len(row), max_col))

    def get_colname_index(self, colname, colnames):
        for index, cname in enumerate(colnames):
            if str(cname) == str(colname):
                return index + 1

    def update_row_by_colname(self, colname, stringrow):
        colnames = self.get_first_column_names()
        index = self.get_colname_index(colname, colnames)
        if index:
            self.update_row_by_rownum(index, stringrow)
        else:
            print("%s not in excel" %colname) 

    def get_col_id_by_name(self, col_name):
        for index, colname in enumerate(self.head_row):
            if col_name == colname.value:
                return index

    def get_sheet_id_by_name(self, sheet_name):
        for index, sheetname in enumerate(self.wb.sheetnames):
            if sheetname == sheet_name:
                return index

    def get_all_json(self):
        self.rows = list(self.sheet.iter_rows())
        row_list = []
        for row in self.rows[1:]:
            row_list.append(dict(zip([cell.value for cell in self.head_row],
                                     [cell.value for cell in row])))
        return row_list

    def insert_json_rows(self, json_rows):
        check_head_row = False
        if not self.head_row:
            self.head_row = json_rows[0].keys()
            self.sheet.append(self.head_row)
            check_head_row = True
        for json_row in json_rows:
            if check_head_row:
                row = [json_row.get(colname) for colname in self.head_row]
            else:
                row = [json_row.get(colname.value) for colname in self.head_row]
            self.sheet.append(row)

    def delete_col(self, colname):
        col_id = self.get_col_id_by_name(colname)
        self.sheet.delete_cols(col_id+1)
